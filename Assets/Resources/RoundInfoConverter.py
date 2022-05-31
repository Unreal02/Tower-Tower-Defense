from xml.etree.ElementTree import tostring
import openpyxl
import json

wb = openpyxl.load_workbook('./RoundInfo.xlsx')
ws = wb['Sheet1']

raw_data = [[]]

cur_round = 0
for r in range(2, ws.max_row + 1):
    round = ws.cell(r, 1).value
    time = ws.cell(r, 2).value
    enemyID = ws.cell(r, 3).value
    if round != None:
        cur_round = round
        raw_data.append([])
    raw_data[cur_round].append((time, enemyID))
print(raw_data)
print()

data = {'rounds': []}
for (round, spawn_list) in enumerate(raw_data):
    if round == 0:
        continue
    spawns = {'spawns': []}
    for spawn in spawn_list:
        spawns['spawns'].append({'time': spawn[0], 'enemy': "Enemy" + str(spawn[1])})
    data['rounds'].append(spawns)
print(data)

with open('./RoundInfo.json', 'w') as file:
    json.dump(data, file, indent = 4)